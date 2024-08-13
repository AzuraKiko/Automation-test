import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from datetime import datetime

def read_test_results(file_path: str) -> pd.DataFrame:
    """
    đọc kết quả trong file test_results.text để clean và tạo pandas DataFrame
    gồm test_case_id và result (PASS, FAIL)
    """
    result_list = open(file_path, "r")

    rows = []

    for line in result_list:
        test_case_id = line.strip().split(":")[0][1:]
        test_case_result = line.strip().split(",")[1].strip()[1:-1]
        rows.append(
            {"test_case_id": test_case_id, "automation_result": test_case_result}
        )

    result_list.close()

    return pd.DataFrame.from_records(rows)




def read_excel_report(file_path: str, column_to_update: str, sheet_name: str) -> pd.DataFrame:
    """
    đọc file excel report và trích xuất cột test_case_id và cột result cần update với automation test cases
    """

    df = pd.read_excel(
        file_path,
        skiprows=20,        # skip những dòng tạo ra format file - chỉ read data
        sheet_name=sheet_name,
    ).iloc[:, [0, int(column_to_update)]]  # lọc columns để lấy test_case_id và cột cần update

    old_names = df.columns
    new_columns_name = ["test_case_id", "current_result"]
    df.rename(columns=dict(zip(old_names, new_columns_name)), inplace=True)

    return df




def update_report_and_save(input_text_file: str, input_excel_file: str, column_to_update: str, sheet_name: str):
    """
    Update file test report với kết quả chạy automation test --> save vào file excel
    """

    test_result: pd.DataFrame = read_test_results(input_text_file)
    excel_report_data: pd.DataFrame = read_excel_report(input_excel_file, column_to_update, sheet_name)

    updated_test_result = excel_report_data.merge(test_result, how="left", on="test_case_id")

    updated_test_result["final_update"] = updated_test_result.apply(
        lambda row: (
            row["automation_result"]
            if (pd.isnull(row["current_result"]))
            else row["current_result"]
        ),
        axis=1,
    )

    # chỉ chọn cột test result để update
    updated_result = updated_test_result[["final_update"]]
    # print(updated_result)


    wb = load_workbook(input_excel_file)
    ws = wb[sheet_name]

    data_rows = dataframe_to_rows(updated_result, index=False, header=False)

    for r_idx, row in enumerate(data_rows, 22):
        for c_idx, value in enumerate(row, int(column_to_update) + 1):
            ws.cell(row=r_idx, column=c_idx, value=value)


    # Không muốn update trực tiếp vào file
    current_time = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
    output_path = input_excel_file.split(".xlsx")[0] + f"_{current_time}" + "_updated.xlsx" 
    wb.save(output_path)




# test end to end flow

# input_excel_file = "/home/nghiaht/workspace/rd-automation-framework/results_report/Portfolio_Testcase_sample_v0.1.xlsx"
# column_to_update = "8"
# input_text_file = "/home/nghiaht/workspace/rd-automation-framework/test_results.txt"

# sheet_name = "Chức năng 1"


# update_report_and_save(input_text_file, input_excel_file, column_to_update, sheet_name)